import json
import pandas as pd
import scraper
from openpyxl import load_workbook, Workbook
import numpy as np
import os

outputDir = "data/"

def scrapeFrom(target : str, returnAvg = True): 
    with open("config/webConfig.json") as f:
        data = json.load(f)
    
    websites = list(data['Websites'].items())
    dataList = []
    
    for i in range(len(data['Websites'])):
        itemPrice = websites[i][1]['ItemPrice']
        print(itemPrice)
        print(websites[i][1]['Url'])
        
        dataList = {"url" : websites[i][0], "data" : scraper.scrapePrices(websites[i][1]['Url'], target, itemPrice['Attribute'],
                             itemPrice['PriceTag'], itemPrice['PriceClass'], 
                             itemPrice['NameTag'], itemPrice['NameClass'], False)}
        
    return dataList

def avg(dataList, target : str, blacklist):
    prices = [item['price'] for item in dataList['data']]
    names = [item['name'] for item in dataList['data']]
    print(prices, names)
    return scraper.format(prices, names, target, blacklist=blacklist)

def toSheet(fileName : str, data, itemName : str):
    path = f"data\\{fileName}.xlsx"
    if not os.path.exists(path):
        wb = Workbook()
    else:
        wb = load_workbook(path)
       
    values = [np.min(data), np.average(data), np.max(data)]    
    try:
        sheet = wb['prices']
    except:
        sheet = wb.create_sheet(title="prices", index=0)
    
        
    #Define columns
    sheet['A1'] = "Item"
    sheet['B1'] = "Price-Optimist"
    sheet['C1'] = "Price-Average"
    sheet['D1'] = "Price-Pessimist"
        
    #Look for empty set of 3 cells to write on
    for i in range(100):
        row = i + 1
        if sheet[f'A{row}'].value: continue
        else:
            sheet[f'A{row}'] = itemName
            for e in range(len(values)):
                sheet[f'{chr(65 + e + 1)}{row}'] = values[e]
            break
        
    wb.save(path)