import scraper as scr
import numpy as np
import override
from openpyxl import load_workbook

def main():
    wb = load_workbook('data\Microcontroler-Datasheet.xlsx')
    
    data = scr.scrapePrices(wb['Specs']['A6'].value, False)
    sheetName = wb['Specs']['A6'].value
    if sheetName in wb:
        sheet = wb[sheetName]
    else: 
        sheet = wb.create_sheet(title=sheetName, index=0)
        
    for i, item in enumerate(data, start=1):
        sheet[f'A{i}'] = item['name']
        sheet[f'B{i}'] = item['price']
        
    data = scr.scrapePrices(wb['Specs']['A7'].value, False)
    sheetName = wb['Specs']['A7'].value
    if sheetName in wb:
        sheet = wb[sheetName]
    else: 
        sheet = wb.create_sheet(title=sheetName, index=0)
        
    for i, item in enumerate(data, start=1):
        sheet[f'A{i}'] = item['name']
        sheet[f'B{i}'] = item['price']
        
        
    wb.save('data\Microcontroler-Datasheet.xlsx')

if __name__ == "__main__":
    main()